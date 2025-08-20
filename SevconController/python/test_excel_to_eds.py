#!/usr/bin/env python3

import os
from openpyxl.cell.cell import MergedCell
from openpyxl import load_workbook
from enum import Enum


class ObjectType(Enum):
    VAR = 7
    ARRAY = 8
    RECORD = 9


class DataType(Enum):
    BOOLEAN = 0
    INTEGER8 = 1
    INTEGER16 = 2
    INTEGER32 = 3
    INTEGER64 = 4
    UNSIGNED8 = 5
    UNSIGNED16 = 6
    UNSIGNED32 = 7
    UNSIGNED64 = 8
    REAL32 = 9
    REAL64 = 10
    VISIBLE_STRING = 11
    OCTET_STRING = 12
    UNICODE_STRING = 13
    DOMAIN = 14


class dictOb():

    def __init__(self, header_dict):
        self.primaryDict = header_dict.copy()
        self.subindex = []

    @property
    def name(self):
        return self.primaryDict.get("Name", None)

    @property
    def index(self):
        return self.primaryDict.get("Index", None)

    @property
    def DataType(self):
        return self.primaryDict.get("Data Type", None)

    @property
    def ObjectType(self):
        if len(self.subindex) > 1:
            # For arrays/records, look at the data type to determine object type
            data_type = self.DataType
            if data_type and "ARRAY" in data_type.upper():
                return "ARRAY"
            elif data_type and "RECORD" in data_type.upper():
                return "RECORD"
            else:
                return "ARRAY"  # Default for multi-subindex objects
        else:
            return "VAR"


def fix_value(this_value):
    # Return if None
    if this_value is None:
        return this_value
    elif this_value == "None":
        return None
    # Return if not a string
    elif not isinstance(this_value, str):
        return this_value
    # Check if string is numberic
    elif this_value.isnumeric():
        return int(this_value)
    # Check if string is hex
    elif this_value[-1:] == 'h':
        try:
            int("0x"+this_value[:-1], 0)
        except:
            return this_value
        return int("0x"+this_value[:-1], 0)
    else:
        return this_value


# Test with the existing Excel file
excel_file = "/mnt/c/REPOS/Voltworks_Garage/Firmware/SevconController/Master Object Dictionary Database.xlsx"

if not os.path.exists(excel_file):
    print(f"Excel file not found: {excel_file}")
    exit(1)

try:
    wb1 = load_workbook(excel_file)
except Exception as e:
    print(f"Error loading Excel file: {e}")
    exit(1)

if 'Object Dictionary' not in wb1.sheetnames:
    print("Error: 'Object Dictionary' sheet not found in the Excel file.")
    print(f"Available sheets: {wb1.sheetnames}")
    exit(1)

wb1.active = wb1['Object Dictionary']

print(f"Excel file loaded successfully. Sheet dimensions: {wb1.active.max_row} x {wb1.active.max_column}")

# Check header row
header_dict = {}
header_row = wb1.active[4]
for i in range(1, 18):
    header_value = header_row[i].value
    if header_value is None:
        break  # Stop if we reach an empty header cell
    header_dict[str(header_value)] = None

print(f"Found {len(header_dict)} headers: {list(header_dict.keys())}")

# Test a few sample rows
sample_count = 0
for row_num, row in enumerate(wb1.active.iter_rows(min_row=5), start=5):
    if sample_count >= 5:  # Only test first 5 rows
        break
        
    # Check if row is empty
    if all(cell.value is None for cell in row):
        continue
    
    sample_count += 1
    print(f"\nRow {row_num}:")
    
    # Convert the row into a dictionary
    tempdict = header_dict.copy()
    for i, item in enumerate(tempdict):
        if i+1 < len(row):
            value = row[i+1].value
            value = fix_value(value)
            tempdict[item] = value
            if value is not None:
                print(f"  {item}: {value}")

print("\nTest completed successfully!")