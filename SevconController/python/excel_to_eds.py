import os
from openpyxl.cell.cell import MergedCell
from openpyxl import load_workbook
from enum import Enum
from datetime import datetime


class ObjectType(Enum):
    VAR = 7
    ARRAY = 8
    RECORD = 9


class DataType(Enum):
    BOOLEAN = 0x01
    INTEGER8 = 0x02
    INTEGER16 = 0x03
    INTEGER32 = 0x04
    UNSIGNED8 = 0x05
    UNSIGNED16 = 0x06
    UNSIGNED32 = 0x07
    REAL32 = 0x08
    VISIBLE_STRING = 0x09
    OCTET_STRING = 0x0A
    UNICODE_STRING = 0x0B
    TIME_OF_DAY = 0x0C
    TIME_DIFFERENCE = 0x0D
    DOMAIN = 0x0F
    INTEGER24 = 0x10
    REAL64 = 0x11
    INTEGER40 = 0x12
    INTEGER48 = 0x13
    INTEGER56 = 0x14
    INTEGER64 = 0x15
    UNSIGNED24 = 0x16
    UNSIGNED40 = 0x18
    UNSIGNED48 = 0x19
    UNSIGNED56 = 0x1A
    UNSIGNED64 = 0x1B


class dictOb():

    def __init__(self, header_dict):
        self.primaryDict = header_dict.copy()
        self.subindex = []

    @property
    def name(self):
        return self.primaryDict.get("Name", None)

    @property
    def index(self):
        return self.primaryDict.get("Index", None)

    @property
    def DataType(self):
        return self.primaryDict.get("Data Type", None)

    @property
    def ObjectType(self):
        if len(self.subindex) > 1:
            # For arrays/records, look at the data type to determine object type
            data_type = self.DataType
            if data_type and "ARRAY" in data_type.upper():
                return "ARRAY"
            elif data_type and "RECORD" in data_type.upper():
                return "RECORD"
            else:
                return "ARRAY"  # Default for multi-subindex objects
        else:
            return "VAR"






EDS_object_format = {
    "SubNumber": None,
    "ObjectType": None,
    "ParameterName": None
}

EDS_sub_format = {
    "ParameterName": None,
    "ParameterValue": None,
    "ObjectType": None,
    "DataType": None,
    "AccessType": None,
    "LowLimit": None,
    "HighLimit": None,
    "PDOMapping": None,
    "DefaultValue": None
}






def format_hex_value(value, digits=4):
    """Format a numeric value as hex with specified digits"""
    if value is None or value == "" or value == "No":
        return ""
    if isinstance(value, str) and value.upper() == "NONE":
        return ""
    try:
        if isinstance(value, str) and value.endswith('h'):
            # Already hex format, convert to int first
            val = int("0x" + value[:-1], 0)
        elif isinstance(value, (int, float)):
            val = int(value)
        else:
            return str(value)
        return f"0x{val:0{digits}X}"
    except:
        return str(value) if value is not None else ""

def write_eds_object(object: dictOb) -> str:
    string = ""
    if len(object.subindex) > 1:
        string += "[{}]\n".format(hex(object.index)[2:])
        string += "{}={}\n".format("ParameterName", object.name)
        
        # Handle ObjectType conversion with error handling
        try:
            object_type_name = object.DataType.upper().replace(" ", "_")
            object_type_val = ObjectType[object_type_name].value
            string += "{}={}\n".format("ObjectType", object_type_val)
        except KeyError:
            print(f"Warning: Unknown object type '{object.DataType}' for object {hex(object.index) if isinstance(object.index, int) else object.index}. Using ARRAY (8).")
            string += "{}={}\n".format("ObjectType", ObjectType.ARRAY.value)
        except Exception as e:
            print(f"Error: Failed to process object type '{object.DataType}' for object {hex(object.index) if isinstance(object.index, int) else object.index}: {e}")
            string += "{}={}\n".format("ObjectType", ObjectType.ARRAY.value)
            
        string += ";StorageLocation=RAM\n"
        string += "{}={}\n\n".format("SubNumber", format_hex_value(len(object.subindex), 1))

    for item in object.subindex:
        if len(object.subindex) == 1:
            string += "[{}]\n".format(hex(object.index)[2:])
        else:
            string += "[{}sub{}]\n".format(hex(object.index)[2:], item["Sub-Index"])
        string += "{}={}\n".format("ParameterName", item["Name"])
        string += "{}={}\n".format("ObjectType", "0x7")
        string += ";StorageLocation=RAM\n"
        
        # Format DataType as 4-digit hex
        try:
            data_type_name = item["Data Type"].upper().replace(" ", "_")
            data_type_val = DataType[data_type_name].value
            string += "{}={}\n".format("DataType", format_hex_value(data_type_val, 4))
        except KeyError as e:
            print(f"Warning: Unknown data type '{item['Data Type']}' for object {hex(object.index) if isinstance(object.index, int) else object.index}, sub-index {item.get('Sub-Index', 0)}. Using 0x0000.")
            string += "{}={}\n".format("DataType", "0x0000")
        except Exception as e:
            print(f"Error: Failed to process data type '{item.get('Data Type', 'Unknown')}' for object {hex(object.index) if isinstance(object.index, int) else object.index}: {e}")
            string += "{}={}\n".format("DataType", "0x0000")
        
        string += "{}={}\n".format("AccessType", item["Access Type"].lower())
        
        # Format limits as hex
        string += "{}={}\n".format("HighLimit", format_hex_value(item["High Limit"]))
        string += "{}={}\n".format("LowLimit", format_hex_value(item["Low Limit"]))
        
        # Format default value  
        value = item.get("Value", "")
        if isinstance(value, int):
            value = format_hex_value(value)
        elif value == "No" or value == "None":
            value = ""
        string += "{}={}\n".format("DefaultValue", value)
        
        # PDO Mapping
        pdo = 0
        if item["Map to PDO?"] == "Yes":
            pdo = 1
        string += "{}={}\n".format("PDOMapping", pdo)
        
        # Add Object Flags if available
        obj_flags = item.get("Object Flags")
        if obj_flags is not None and obj_flags != "" and obj_flags != "No":
            string += "{}={}\n".format("ObjFlags", obj_flags)
            
        string += "\n"

    return string

def fix_value(this_value):
    # Return if None
    if this_value is None:
        return this_value
    elif this_value == "None":
        return None
    # Return if not a string
    elif not isinstance(this_value, str):
        return this_value
    # Check if string is numberic
    elif this_value.isnumeric():
        return int(this_value)
    # Check if string is hex
    elif this_value[-1:] == 'h':
        try:
            int("0x"+this_value[:-1], 0)
        except:
            return this_value
        return int("0x"+this_value[:-1], 0)
    else:
        return this_value


file1 = "../Master Object Dictionary Database.xlsx"

try:
    wb1 = load_workbook(file1)
except Exception as e:
    print(f"Error loading Excel file: {e}")
    exit()

if 'Object Dictionary' not in wb1.sheetnames:
    print("Error: 'Object Dictionary' sheet not found in the Excel file.")
    print(f"Available sheets: {wb1.sheetnames}")
    exit()

wb1.active = wb1['Object Dictionary']

header_dict = {}
header_row = wb1.active[4]
for i in range(1, 18):
    header_value = header_row[i].value
    if header_value is None:
        break  # Stop if we reach an empty header cell
    header_dict[str(header_value)] = None

if len(header_dict) == 0:
    print("Error: No valid headers found in row 4")
    exit()

print(f"Found {len(header_dict)} headers: {list(header_dict.keys())}")

Object_Dictionary = {}
newOb = None
index = 0

for row_num, row in enumerate(wb1.active.iter_rows(min_row=5), start=5):
    # Check if row is empty
    if all(cell.value is None for cell in row):
        continue
        
    # Convert the row into a dictionary
    tempdict = header_dict.copy()
    for i, item in enumerate(tempdict):
        if i+1 < len(row):
            value = row[i+1].value
            value = fix_value(value)
            tempdict[item] = value

    # Skip rows without an index or with invalid data, unless it's a merged cell (subindex)
    if tempdict.get("Index") is None and not isinstance(row[1], MergedCell):
        continue
    
    # If we encounter a new object (with non-None index), finalize the previous object first
    if tempdict.get("Index") is not None and newOb is not None:
        # Finalize and save the previous object
        if newOb.index is not None and newOb.index != "None":
            Object_Dictionary[newOb.index] = newOb
            print(f"Added object {hex(newOb.index) if isinstance(newOb.index, int) else newOb.index}")
        newOb = None
        
    # Check for single variable object
    if (tempdict["Index"] is not None) and (tempdict["Sub-Index"] == 0):
        newOb = dictOb(header_dict)
        newOb.primaryDict = tempdict.copy()
        newOb.subindex.append(tempdict.copy())

    # Check for array/record object (Sub-Index is None)
    elif tempdict["Sub-Index"] is None and tempdict.get("Index") is not None:
        newOb = dictOb(header_dict)
        newOb.primaryDict = tempdict.copy()
        index = 0

    # Check for sub-index objects (merged cell)
    elif isinstance(row[1], MergedCell):
        if newOb is None:
            print(f"Warning: Sub-index found without parent object at row {row_num}, skipping...")
            continue
        tempdict["Index"] = newOb.primaryDict["Index"]
        
        # Check for duplicate sub-indices and skip them
        current_sub_index = tempdict["Sub-Index"]
        duplicate_found = False
        for existing_item in newOb.subindex:
            if existing_item.get("Sub-Index") == current_sub_index:
                print(f"Warning: Duplicate sub-index {current_sub_index} found at row {row_num} for object {hex(newOb.index) if isinstance(newOb.index, int) else newOb.index}, skipping...")
                duplicate_found = True
                break
        
        if not duplicate_found:
            newOb.subindex.append(tempdict.copy())
            index += 1

    # Handle other cases  
    else:
        # Provide more detailed info about unidentified structures
        index_val = tempdict.get("Index")
        sub_index_val = tempdict.get("Sub-Index")
        name_val = tempdict.get("Name", "Unknown")
        print(f"Warning: Unidentified object structure at row {row_num} - Index: {index_val}, Sub-Index: {sub_index_val}, Name: '{name_val}'. Treating as single variable.")
        
        if index_val is not None:
            newOb = dictOb(header_dict)
            newOb.primaryDict = tempdict.copy()
            newOb.subindex.append(tempdict.copy())
        else:
            print(f"Error: Skipping row {row_num} - no valid index found")

# Don't forget to add the last object
if newOb is not None and newOb.index is not None and newOb.index != "None":
    Object_Dictionary[newOb.index] = newOb
    print(f"Added object {hex(newOb.index) if isinstance(newOb.index, int) else newOb.index}")
elif newOb is not None:
    print(f"Warning: Final object not added - invalid index: {newOb.index}")

print(f"Processing complete - {len(Object_Dictionary)} objects loaded")

# Print summary statistics
mandatory_count = 0
communication_count = 0
manufacturer_count = 0
device_count = 0

for obj_index in Object_Dictionary:
    if obj_index in [0x1000, 0x1001, 0x1018]:
        mandatory_count += 1
    elif 0x1000 <= obj_index <= 0x1FFF:
        communication_count += 1
    elif 0x2000 <= obj_index <= 0x5FFF:
        manufacturer_count += 1
    elif 0x6000 <= obj_index <= 0x9FFF:
        device_count += 1

print(f"Object distribution: Mandatory={mandatory_count}, Communication={communication_count}, Manufacturer={manufacturer_count}, Device Profile={device_count}")

path = os.path.dirname(file1)

with open(os.path.join(path, 'new_EDS.eds'), 'w') as f:
    # Write EDS header sections
    f.write("[FileInfo]\n")
    f.write("FileName=new_EDS.eds\n")
    f.write("FileVersion=1\n")
    f.write("FileRevision=0\n")
    f.write("LastEDS=\n")
    f.write("EDSVersion=4.0\n")
    f.write("Description=Generated from Excel Object Dictionary\n")
    
    # Generate DS306 compliant timestamps
    now = datetime.now()
    creation_time = now.strftime("%I:%M%p")  # 12-hour format with AM/PM
    creation_date = now.strftime("%m-%d-%Y")  # MM-DD-YYYY format
    
    f.write(f"CreationTime={creation_time}\n")
    f.write(f"CreationDate={creation_date}\n")
    f.write("CreatedBy=Python Script\n")
    f.write(f"ModificationTime={creation_time}\n")
    f.write(f"ModificationDate={creation_date}\n")
    f.write("ModifiedBy=Python Script\n")
    f.write("\n")
    
    f.write("[DeviceInfo]\n")
    f.write("VendorName=Tech/Ops SEVCON Ltd\n")
    f.write("VendorNumber=0x0000001e\n")
    f.write("ProductName=Gen4\n")
    f.write("ProductNumber=0x07065037\n")
    f.write("RevisionNumber=65559\n")
    f.write("BaudRate_10=0\n")
    f.write("BaudRate_20=0\n")
    f.write("BaudRate_50=0\n")
    f.write("BaudRate_125=0\n")
    f.write("BaudRate_250=0\n")
    f.write("BaudRate_500=1\n")
    f.write("BaudRate_800=0\n")
    f.write("BaudRate_1000=0\n")
    f.write("SimpleBootUpMaster=1\n")
    f.write("SimpleBootUpSlave=1\n")
    f.write("Granularity=64\n")
    f.write("DynamicChannelsSupported=0\n")
    f.write("CompactPDO=0\n")
    f.write("GroupMessaging=0\n")
    f.write("NrOfRXPDO=5\n")
    f.write("NrOfTXPDO=5\n")
    f.write("LSS_Supported=0\n")
    f.write("\n")
    
    f.write("[DummyUsage]\n")
    f.write("Dummy0001=0\n")
    f.write("Dummy0002=0\n")
    f.write("Dummy0003=0\n")
    f.write("Dummy0004=0\n")
    f.write("Dummy0005=0\n")
    f.write("Dummy0006=0\n")
    f.write("Dummy0007=0\n")
    f.write("\n")
    
    f.write("[Comments]\n")
    f.write("Lines=1\n")
    f.write("Line1=Generated from Excel Object Dictionary\n")
    f.write("\n")
    
    # Object sections start here

    Mandatory_object_area = [0x1000, 0x1001, 0x1018]
    Communication_object_area = range(0x1000, 0x1FFF)
    Manufacturer_object_area = range(0x2000, 0x5FFF)
    Device_object_area = range(0x6000, 0x9FFF)
    rpdo_comm_range = range(0x1400, 0x1600)      # RPDO Communication Parameters
    rpdo_mapping_range = range(0x1600, 0x1800)  # RPDO Mapping Parameters
    tpdo_comm_range = range(0x1800, 0x1A00)     # TPDO Communication Parameters  
    tpdo_mapping_range = range(0x1A00, 0x1C00)  # TPDO Mapping Parameters

    range_comm = {}
    range_mfg = {}
    range_dev = {}
    ran_man = {}



    for object in Object_Dictionary:
        obj_index = Object_Dictionary[object].index
        
        if obj_index in Mandatory_object_area:
            ran_man[object] = Object_Dictionary[object]
        elif (obj_index in Communication_object_area or 
              obj_index in Device_object_area or
              obj_index in rpdo_comm_range or 
              obj_index in rpdo_mapping_range or
              obj_index in tpdo_comm_range or 
              obj_index in tpdo_mapping_range):
            range_comm[object] = Object_Dictionary[object]
        elif obj_index in Manufacturer_object_area:
            range_mfg[object] = Object_Dictionary[object]
        else:
            print(f'Warning: Object {hex(object)} out of standard ranges, placing in manufacturer area')
            range_mfg[object] = Object_Dictionary[object]

    f.write("[MandatoryObjects]\nSupportedObjects={}\n".format(len(ran_man)))
    for num, object in enumerate(ran_man):
        f.write("{}={}\n".format(num+1, hex(object).upper()))
    f.write("\n")
    for object in ran_man:
        f.write(write_eds_object(Object_Dictionary[object]))

    f.write("[OptionObjects]\nSupportedObjects={}\n".format(len(range_comm)))
    for num, object in enumerate(range_comm):
        f.write("{}={}\n".format(num+1, hex(object).upper()))
    f.write("\n")
    for object in range_comm:
        f.write(write_eds_object(Object_Dictionary[object]))

    f.write("[ManufacturerObjects]\nSupportedObjects={}\n".format(len(range_mfg)))
    for num, object in enumerate(range_mfg):
        f.write("{}={}\n".format(num+1, hex(object).upper()))
    f.write("\n")
    for object in range_mfg:
        f.write(write_eds_object(Object_Dictionary[object]))

    # f.write("[OptionObjects]\nSupportedObjects={}\n".format(len(range_comm)))
    # for num, object in enumerate(range_comm):
    #     f.write("{}={}".format(num+1, hex(object)))
    # for object in range_comm:
    #     write_esd_object(Object_Dictionary[object])



print("done")


